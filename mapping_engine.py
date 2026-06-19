import os
import json
import base64
import io
import re
import zlib
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from sqlalchemy.orm import Session
from database import DbFile, CategoryConfig, SizeMapping, ProcessingJob, Setting

def decode_and_decompress(b64_str: str) -> bytes:
    """Decode base64 string and decompress using zlib. Falls back to raw bytes if not compressed."""
    try:
        raw_bytes = base64.b64decode(b64_str)
        try:
            return zlib.decompress(raw_bytes)
        except zlib.error:
            # If it is not compressed (old format), return the raw bytes directly
            return raw_bytes
    except Exception:
        if isinstance(b64_str, str):
            return b64_str.encode("utf-8")
        return b64_str

class EngineLogger:
    def __init__(self, job_id=None, log_queue=None):
        self.job_id = job_id
        self.log_queue = log_queue
        self.logs_list = []

    def log(self, level, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = {
            "time": timestamp,
            "level": level.upper(),
            "message": message
        }
        self.logs_list.append(log_entry)
        print(f"[{level.upper()}] {message}")
        if self.log_queue is not None:
            self.log_queue.put(log_entry)

def normalize_color_val(c):
    if not c or pd.isna(c):
        return ""
    import re
    return re.sub(r'[\s\.\-_]+', '', str(c)).upper()

def map_dropdown_gender(gender_val):
    if not gender_val:
        return "Boys"
    g = str(gender_val).strip().upper()
    if 'GIRL' in g or 'FEMALE' in g:
        return "Girls"
    elif 'BOY' in g or 'MALE' in g:
        return "Boys"
    elif 'UNISEX' in g or 'KID' in g:
        return "Unisex"
    elif 'WOMEN' in g:
        return "Women"
    elif 'MEN' in g:
        return "Men"
    return "Boys" # Default fallback

def get_design_code(category, gender):
    cat = str(category).lower().strip()
    g = str(gender).lower().strip()
    
    girls_first_cats = ["tshirt", "t-shirt", "dress", "top", "kurta", "ethnic", "apparel"]
    is_girls_first = any(keyword in cat for keyword in girls_first_cats)
    
    if is_girls_first:
        if "girl" in g or "female" in g:
            return "Group-1"
        elif "boy" in g or "male" in g:
            return "Group-2"
        else:
            return "Group-1"
    else:
        if "boy" in g or "male" in g:
            return "Group-1"
        elif "girl" in g or "female" in g:
            return "Group-2"
        else:
            return "Group-1"

def get_singular_category_item(category_name):
    cat = str(category_name).lower().strip()
    if "dress" in cat:
        return "Dress"
    elif "tshirt" in cat or "t-shirt" in cat:
        return "Tshirt"
    elif "short" in cat:
        return "Shorts"
    elif "trouser" in cat or "jean" in cat or "pant" in cat:
        return "Trouser"
    elif "shoe" in cat or "sandal" in cat or "footwear" in cat:
        return "Pair of Shoes"
    elif "top" in cat:
        return "Top"
    res = category_name.strip()
    if res.endswith("s") and not res.lower().endswith("dress") and not res.lower().endswith("shorts"):
        res = res[:-1]
    return res

def map_size_to_dropdown_age(size_str, options):
    if not size_str:
        return ""
    s = str(size_str).strip().upper()
    if s == 'NB':
        return 'Newborn'
        
    import re
    nums = re.findall(r'\d+\.?\d*', s)
    if not nums:
        return size_str
        
    is_months = 'M' in s and 'Y' not in s
    
    # Try finding exact number range with Years/Months
    for opt in options:
        opt_upper = opt.upper()
        opt_nums = re.findall(r'\d+\.?\d*', opt_upper)
        if opt_nums == nums:
            opt_is_months = 'MONTH' in opt_upper
            if is_months == opt_is_months:
                return opt
                
    # Fallback to general replacement
    if len(nums) == 2:
        unit = 'Months' if is_months else 'Years'
        formatted = f'{nums[0]}-{nums[1]} {unit}'
        for opt in options:
            if opt.lower() == formatted.lower():
                return opt
        return formatted
    elif len(nums) == 1:
        unit = 'Months' if is_months else 'Years'
        formatted = f'{nums[0]} {unit}'
        for opt in options:
            if opt.lower() == formatted.lower():
                return opt
        return formatted
        
    return size_str

# Sizing Charts Global Lookup Databases
GIRLS_SIZE_CHARTS = {}
BOYS_SIZE_CHARTS = {}

def parse_apparel_sizes_file(filepath):
    import pandas as pd
    df = pd.read_excel(filepath, header=None)
    sections = {}
    
    current_section = None
    sizes = []
    
    row_idx = 0
    while row_idx < len(df):
        row = [x if pd.notna(x) else None for x in df.iloc[row_idx].tolist()]
        if all(x is None for x in row):
            row_idx += 1
            continue
            
        first_val = str(row[0]).strip() if row[0] is not None else ""
        
        # Check size header row (contains POM or CM)
        if 'POM' in first_val.upper() or 'CM' in first_val.upper():
            sizes = [str(x).strip() for x in row[1:] if x is not None]
            row_idx += 1
            continue
            
        # Check section header (first column non-null, rest are None)
        if row[0] is not None and all(x is None for x in row[1:]):
            current_section = first_val.upper()
            sections[current_section] = {}
            sizes = []
            row_idx += 1
            continue
            
        # Measurement row
        if current_section and sizes and row[0] is not None:
            meas_name = first_val.upper()
            sections[current_section][meas_name] = {}
            for col_idx, sz in enumerate(sizes):
                if col_idx + 1 < len(row):
                    val = row[col_idx + 1]
                    if val is not None:
                        try:
                            sections[current_section][meas_name][sz] = float(val)
                        except ValueError:
                            sections[current_section][meas_name][sz] = val
        row_idx += 1
    return sections

def load_apparel_size_charts():
    global GIRLS_SIZE_CHARTS, BOYS_SIZE_CHARTS
    girls_path = "Raw Files/Girls Apperel Sizes.xlsx"
    boys_path = "Raw Files/Boys Apparel Sizes.xlsx"
    
    if os.path.exists(girls_path):
        try:
            GIRLS_SIZE_CHARTS = parse_apparel_sizes_file(girls_path)
            print(f"Successfully loaded Girls size charts section headers.")
        except Exception as e:
            print(f"Failed to load Girls size charts: {e}")
            
    if os.path.exists(boys_path):
        try:
            BOYS_SIZE_CHARTS = parse_apparel_sizes_file(boys_path)
            print(f"Successfully loaded Boys size charts section headers.")
        except Exception as e:
            print(f"Failed to load Boys size charts: {e}")

def normalize_size_for_lookup(sz):
    s = str(sz).strip().upper()
    s = s.replace("YRS", "Y").replace("YEARS", "Y")
    s = s.replace("MONTHS", "M").replace("MONTH", "M")
    if 'M' in s:
        s = s.replace("/", "-")
    elif 'Y' in s:
        s = s.replace("-", "/")
    return s

def get_size_chart_section(category, gender):
    cat = str(category).lower().strip()
    g = str(gender).lower().strip()
    
    is_girls = "girl" in g or "female" in g
    
    if is_girls:
        if "tshirt" in cat or "t-shirt" in cat:
            return "GIRLS  T-SHIRT"
        elif "shirt" in cat:
            return "GIRLS SHIRT"
        elif "dress" in cat:
            return "GIRLS  DRESS"
        elif "dungaree" in cat:
            if "dress" in cat:
                return "DRESS LENGTH DUNGAREE"
            return "GIRLS FULL LENGTH DUNGAREE"
        elif "jumpsuit" in cat:
            return "GIRLS  JUMPSUIT"
        elif "sweatshirt" in cat:
            return "GIRLS SWEATSHIRT"
        elif "jeans" in cat or "trouser" in cat or "pant" in cat:
            return "GIRLS JEANS / TROUSER"
        elif "jagging" in cat or "jegging" in cat:
            return "GIRLS JAGGING"
        elif "shorts" in cat or "short" in cat:
            return "GIRLS SHORTS"
        elif "skirt" in cat:
            return "GIRLS SKIRT"
        elif "jacket" in cat:
            return "GIRLS JACKET"
        elif "set" in cat or "clothing set" in cat:
            return "CLOTHING SET"
        if "bottom" in cat or "legging" in cat:
            return "GIRLS JEANS / TROUSER"
        return "GIRLS  TOP"
    else:
        if "bermuda" in cat:
            return "BOYS BERMUDA"
        elif "tshirt" in cat or "t-shirt" in cat:
            return "BOYS T-SHIRT"
        elif "shirt" in cat:
            return "BOYS SHIRT"
        elif "sweatshirt" in cat:
            return "BOYS SWEATSHIRT"
        elif "jacket" in cat:
            return "BOYS JACKET"
        elif "jeans" in cat or "trouser" in cat or "pant" in cat:
            return "BOYS JEANS / TROUSER"
        elif "lower" in cat or "jogger" in cat:
            return "BOYS LOWER /JOGGER"
        elif "shorts" in cat or "short" in cat:
            return "BOYS SHORTS"
        if "bottom" in cat:
            return "BOYS JEANS / TROUSER"
        return "BOYS T-SHIRT"

def get_chart_measurement_by_keyword(gender, section, col_name, brand_size):
    global GIRLS_SIZE_CHARTS, BOYS_SIZE_CHARTS
    g = str(gender).lower().strip()
    is_girls = "girl" in g or "female" in g
    
    charts = GIRLS_SIZE_CHARTS if is_girls else BOYS_SIZE_CHARTS
    
    if not charts or section not in charts:
        return None
        
    sec_data = charts[section]
    
    # Identify keyword to search in size chart row names
    col_upper = str(col_name).strip().upper()
    keyword = None
    if "CHEST" in col_upper or "BUST" in col_upper:
        keyword = "CHEST"
    elif "WAIST" in col_upper:
        keyword = "WAIST"
    elif "HIP" in col_upper:
        keyword = "HIP"
    elif "SHOULDER" in col_upper:
        keyword = "SHOULDER"
    elif "INSEAM" in col_upper:
        keyword = "INSEAM"
    elif "SLEEVE" in col_upper:
        keyword = "SLEEVE"
    elif "BOTTOM" in col_upper and "LENGTH" in col_upper:
        keyword = "LENGTH"
    elif "LENGTH" in col_upper:
        keyword = "LENGTH"
        
    if not keyword:
        return None
        
    is_bottoms_section = any(x in section for x in ["JEANS", "TROUSER", "LOWER", "JOGGER", "SHORTS", "BERMUDA", "DUNGAREE", "SKIRT", "JAGGING"])
    
    # Find matching measurement row key
    target_key = None
    for k in sec_data.keys():
        k_upper = str(k).strip().upper()
        if keyword in k_upper:
            # Special check for bottoms length / inseam
            if "BOTTOM" in col_upper or "INSEAM" in col_upper:
                if is_bottoms_section or any(x in k_upper for x in ["LOWER", "BOTTOM", "LOWER WEAR", "INSEAM", "TROUSER", "PANT"]):
                    target_key = k
                    break
            else:
                if is_bottoms_section and keyword == "LENGTH" and not any(x in col_upper for x in ["BOTTOM", "INSEAM"]):
                    continue
                target_key = k
                if not any(x in k_upper for x in ["LOWER", "BOTTOM", "LOWER WEAR", "INSEAM"]):
                    break
                
    if not target_key:
        return None
        
    meas_data = sec_data[target_key]
    
    # Normalize size key for lookup
    norm_size = normalize_size_for_lookup(brand_size)
    
    if norm_size in meas_data:
        return meas_data[norm_size]
        
    if str(brand_size).strip() in meas_data:
        return meas_data[str(brand_size).strip()]
        
    for k in meas_data.keys():
        k_str = str(k).strip().upper()
        if k_str == norm_size or k_str == str(brand_size).strip().upper():
            return meas_data[k]
            
    return None

def convert_cm_to_inches(val):
    if val is None:
        return None
    try:
        f_val = float(val)
        return round(f_val / 2.54, 1)
    except (ValueError, TypeError):
        return val

def clean_excel_value(val):
    if pd.isna(val) or val is None:
        return ""
    # Remove trailing .0 from float integers
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def learn_from_historical_excel(db: Session, file_content_b64: str, filename: str, category_name: str, logger: EngineLogger):
    logger.log("info", f"Analyzing historical listing file '{filename}' for category '{category_name}'...")
    
    file_data = decode_and_decompress(file_content_b64)
    wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
    
    # 1. Identify product sheet
    target_sheet_name = None
    for sname in wb.sheetnames:
        if sname not in ["Instructions Sheet", "Instructions", "mastersheet", "masterdata"]:
            target_sheet_name = sname
            break
            
    if not target_sheet_name:
        raise ValueError("Could not find a valid category data sheet in the historical file.")
        
    sheet = wb[target_sheet_name]
    logger.log("info", f"Analyzing sheet '{target_sheet_name}' (total rows: {sheet.max_row})...")
    
    # 2. Row 3 has technical column names
    headers = [str(sheet.cell(row=3, column=c).value).strip() for c in range(1, sheet.max_column + 1)]
    logger.log("info", f"Found {len(headers)} columns in sheet row 3.")
    
    # Check if we have data (starting at Row 4)
    data_rows = []
    for r in range(4, sheet.max_row + 1):
        row_vals = {}
        row_has_data = False
        for col_idx, h in enumerate(headers, start=1):
            if h and h != "None":
                val = sheet.cell(row=r, column=col_idx).value
                row_vals[h] = val
                if val is not None and str(val).strip() != "":
                    row_has_data = True
        if row_has_data:
            data_rows.append(row_vals)
            
    logger.log("info", f"Extracted {len(data_rows)} data rows from historical file.")
    if not data_rows:
        wb.close()
        return
        
    # 3. Detect hardcoded values
    hardcoded = {}
    for h in headers:
        if not h or h == "None":
            continue
        # Check values across all data rows
        vals = [str(r[h]).strip() if r.get(h) is not None else "" for r in data_rows]
        unique_vals = set(vals)
        # If all rows have the exact same non-empty value, it's hardcoded!
        if len(unique_vals) == 1:
            val = list(unique_vals)[0]
            if val != "":
                # Don't hardcode structural identifiers
                if h not in ['Vendor SKU Code', 'Ean Codes', 'brand  size', 'Style Code', 'Product Name', 'Description', 'Price', 'Color', 'Design Code', 'Front Image', 'Back Image', 'Additional Image 1', 'Additional Image 2', 'Additional Image 3', 'Additional Image 4', 'Additional Image 5', 'Additional Image 6', 'Additional Image 7', 'Additional Image 8']:
                    hardcoded[h] = val
                    
    logger.log("success", f"Identified {len(hardcoded)} hardcoded fields: {list(hardcoded.keys())}")
    
    # 4. Extract Size Mappings
    size_mappings_extracted = 0
    # Identify measurement columns (columns that represent dimensions or measurements)
    measurement_cols = []
    for h in headers:
        h_lower = h.lower()
        if any(keyword in h_lower for keyword in ["garment", "body", "inches", "cm", "chest", "bust", "waist", "hip", "shoulder", "length", "sleeve"]):
            if h not in ['brand  size', 'Style Code', 'Vendor SKU Code', 'Ean Codes', 'Description']:
                measurement_cols.append(h)
                
    logger.log("info", f"Identified measurement columns for size charts: {measurement_cols}")
    
    # Group size measurements by size label
    size_charts = {}
    for r in data_rows:
        size_label = str(r.get('brand  size', '')).strip()
        if not size_label or size_label == "None":
            continue
        
        # Build measurements dict
        measure_dict = {}
        for m_col in measurement_cols:
            m_val = r.get(m_col)
            if m_val is not None:
                measure_dict[m_col] = m_val
                
        if measure_dict:
            size_charts[size_label] = measure_dict
            
    # Save sizes to database
    # First, clear existing size mappings for this category
    db.query(SizeMapping).filter(SizeMapping.category_name == category_name).delete()
    for size_label, measurements in size_charts.items():
        sz = SizeMapping(
            category_name=category_name,
            brand_size=size_label,
            measurements=measurements
        )
        db.add(sz)
        size_mappings_extracted += 1
        
    logger.log("success", f"Saved {size_mappings_extracted} size charts in Sizing Center.")
    
    # 5. Core Column Mapping Config
    col_mappings = {
        "Vendor SKU Code": "ITEM CODE",
        "Ean Codes": "ITEM CODE",
        "Style Code": "ITEM NAME",
        "Price": "MRP",
        "Color": "COLOR",
        "brand  size": "SIZE",
        "Country of Origin": "IMPORTED/DOMESTIC",
        "HSN Codes": "HS CODE",
        "Product Name": "Nykaa Title",
        "Description": "Description",
        "Gender": "GENDER",
        "Design Code": "Design Code Group"
    }
    
    # Save category config
    config = db.query(CategoryConfig).filter(CategoryConfig.category_name == category_name).first()
    if not config:
        config = CategoryConfig(category_name=category_name)
        db.add(config)
        
    config.hardcoded_values = hardcoded
    config.column_mappings = col_mappings
    db.commit()
    wb.close()
    logger.log("success", f"Auto-learning completed for '{category_name}'. Configuration saved successfully.")

def get_column_letter(col_idx):
    """Convert a 1-based column index to an Excel column letter."""
    letter = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter

def get_sheet_xml_path(template_bytes, target_sheet_name):
    import zipfile
    import xml.etree.ElementTree as ET
    import io
    
    with zipfile.ZipFile(io.BytesIO(template_bytes), 'r') as z:
        wb_xml = z.read('xl/workbook.xml')
        wb_root = ET.fromstring(wb_xml)
        
        sheets_el = None
        for child in wb_root.iter():
            tag_local = child.tag.split('}')[-1]
            if tag_local == 'sheets':
                sheets_el = child
                break
                
        if sheets_el is None:
            raise ValueError("Could not find 'sheets' element in workbook.xml")
            
        rId = None
        for sheet_el in sheets_el:
            tag_local = sheet_el.tag.split('}')[-1]
            if tag_local == 'sheet':
                name = sheet_el.attrib.get('name')
                if name and name.strip().lower() == target_sheet_name.strip().lower():
                    for attr_key, attr_val in sheet_el.attrib.items():
                        if attr_key.endswith('}id') or attr_key == 'id':
                            rId = attr_val
                            break
                    if rId:
                        break
                        
        if not rId:
            raise ValueError(f"Could not find relation ID for sheet '{target_sheet_name}' in workbook.xml")
            
        rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_root = ET.fromstring(rels_xml)
        
        target_path = None
        for rel in rels_root:
            tag_local = rel.tag.split('}')[-1]
            if tag_local == 'Relationship':
                if rel.attrib.get('Id') == rId:
                    target_path = rel.attrib.get('Target')
                    break
                    
        if not target_path:
            raise ValueError(f"Could not find target path for relation ID '{rId}' in workbook.xml.rels")
            
        if not target_path.startswith('xl/'):
            target_path = 'xl/' + target_path
            
        return target_path

def read_col_mapping_via_openpyxl_read_only(template_bytes, category_name, logger):
    import openpyxl
    import io
    
    col_mapping = {}
    size_columns = []
    age_options = []
    kids_product_type_options = []
    material_options = []
    resolved_sheet_name = None
    
    def is_size_column(col_name):
        col_lower = col_name.lower()
        if "inches" not in col_lower:
            return False
        keywords = ["chest", "bust", "waist", "hip", "shoulder", "inseam", "sleeve", "length"]
        excludes = ["brand  size", "disclaimer", "type", "fit", "pattern", "description",
                    "neckline", "pocket", "style", "character", "collection", "classification"]
        return any(kw in col_lower for kw in keywords) and not any(ex in col_lower for ex in excludes)

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), read_only=True)
    try:
        # Find target sheet
        for sname in wb.sheetnames:
            if sname not in ["Instructions Sheet", "Instructions", "mastersheet", "masterdata"]:
                if sname.lower().strip() == category_name.lower().strip():
                    resolved_sheet_name = sname
                    break
        if not resolved_sheet_name:
            for sname in wb.sheetnames:
                if sname not in ["Instructions Sheet", "Instructions", "mastersheet", "masterdata"]:
                    resolved_sheet_name = sname
                    break
                    
        if not resolved_sheet_name:
            raise ValueError("Could not find a valid category data sheet in template workbook.")
            
        ws = wb[resolved_sheet_name]
        
        # Read Row 3
        row3 = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 3:
                row3 = row
                break
                
        if row3:
            for c_idx, val in enumerate(row3, start=1):
                if val is not None:
                    col_name_str = str(val).strip()
                    if col_name_str:
                        col_mapping[col_name_str] = c_idx
                        if is_size_column(col_name_str):
                            size_columns.append(col_name_str)
                            
        # Read dropdown options from mastersheet
        if "mastersheet" in wb.sheetnames:
            ws_master = wb["mastersheet"]
            age_col_idx = None
            kpt_col_idx = None
            mat_col_idx = None
            
            row3_master = None
            for idx, r in enumerate(ws_master.iter_rows(min_row=3, max_row=3, values_only=True), start=3):
                row3_master = r
                break
                
            if row3_master:
                for c_idx, val in enumerate(row3_master, start=1):
                    if val:
                        val_lower = str(val).strip().lower()
                        if val_lower == "age":
                            age_col_idx = c_idx
                        elif val_lower in ["kids product type", "product type"]:
                            kpt_col_idx = c_idx
                        elif val_lower == "material":
                            mat_col_idx = c_idx
            
            # Fallbacks if headers not matched
            if age_col_idx is None: age_col_idx = 39
            if kpt_col_idx is None: kpt_col_idx = 27
            if mat_col_idx is None: mat_col_idx = 23
            
            # Read columns dynamically
            # 1. Read age
            for idx, row in enumerate(ws_master.iter_rows(min_col=age_col_idx, max_col=age_col_idx, min_row=4, max_row=350, values_only=True), start=4):
                val = row[0]
                if val is not None:
                    val_str = str(val).strip()
                    if val_str: age_options.append(val_str)
                    
            # 2. Read product types
            for idx, row in enumerate(ws_master.iter_rows(min_col=kpt_col_idx, max_col=kpt_col_idx, min_row=4, max_row=150, values_only=True), start=4):
                val = row[0]
                if val is not None:
                    val_str = str(val).strip()
                    if val_str: kids_product_type_options.append(val_str)
                    
            # 3. Read materials
            for idx, row in enumerate(ws_master.iter_rows(min_col=mat_col_idx, max_col=mat_col_idx, min_row=4, max_row=200, values_only=True), start=4):
                val = row[0]
                if val is not None:
                    val_str = str(val).strip()
                    if val_str: material_options.append(val_str)
                        
        logger.log("info", f"Read {len(col_mapping)} columns, {len(age_options)} age options, {len(kids_product_type_options)} product type options, and {len(material_options)} material options from template.")
        return resolved_sheet_name, col_mapping, size_columns, age_options, kids_product_type_options, material_options
    finally:
        wb.close()

def fill_template_via_xml_manipulation(template_bytes, target_sheet_name, col_mapping, rows_data, logger):
    import zipfile
    import io
    import xml.etree.ElementTree as ET
    import re
    
    sheet_xml_path = get_sheet_xml_path(template_bytes, target_sheet_name)
    logger.log("info", f"Mapped sheet '{target_sheet_name}' to XML file '{sheet_xml_path}'")
    
    with zipfile.ZipFile(io.BytesIO(template_bytes), 'r') as zin:
        files = {name: zin.read(name) for name in zin.namelist()}
        
    sheet_xml_bytes = files[sheet_xml_path]
    sheet_xml_str = sheet_xml_bytes.decode('utf-8')
    
    # 1. Locate sheetData tags
    start_tag = "<sheetData"
    end_tag = "</sheetData>"
    
    start_idx = sheet_xml_str.find(start_tag)
    end_idx = sheet_xml_str.find(end_tag)
    
    if start_idx == -1 or end_idx == -1:
        raise ValueError("Could not find sheetData tags")
        
    start_close_idx = sheet_xml_str.find(">", start_idx)
    sheet_data_open_tag = sheet_xml_str[start_idx:start_close_idx + 1]
    
    prefix = sheet_xml_str[:start_idx]
    sheet_data_content = sheet_xml_str[start_close_idx + 1:end_idx]
    suffix = sheet_xml_str[end_idx + len(end_tag):]
    
    # 2. Parse existing sheetData rows
    # Wrap in dummy root with namespaces to parse row 1, 2, 3 safely
    wrapped_content = f'<root xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac">{sheet_data_content}</root>'
    
    root_el = ET.fromstring(wrapped_content.encode('utf-8'))
    
    rows_to_keep = []
    for row in root_el:
        if row.attrib.get('r') in ['1', '2', '3']:
            rows_to_keep.append(row)
            
    # 3. Create and append new rows
    ns_uri = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    
    for i, sku_data in enumerate(rows_data):
        row_num = 4 + i
        row_el = ET.Element(f"{ns_uri}row", {'r': str(row_num)})
        
        row_cells = []
        for col_name, col_num in col_mapping.items():
            val = sku_data.get(col_name)
            if val is not None and val != "":
                row_cells.append((col_num, val))
                
        row_cells.sort(key=lambda x: x[0])
        
        for col_num, val in row_cells:
            col_letter = get_column_letter(col_num)
            cell_ref = f"{col_letter}{row_num}"
            
            c_attrib = {'r': cell_ref}
            is_string = False
            val_str = ""
            if isinstance(val, bool):
                c_attrib['t'] = 'b'
                val_str = '1' if val else '0'
            elif isinstance(val, (int, float)):
                val_str = str(val)
            else:
                is_string = True
                c_attrib['t'] = 'inlineStr'
                val_str = str(val)
                
            c_el = ET.Element(f"{ns_uri}c", c_attrib)
            if is_string:
                is_el = ET.Element(f"{ns_uri}is")
                t_el = ET.Element(f"{ns_uri}t")
                t_el.text = val_str
                is_el.append(t_el)
                c_el.append(is_el)
            else:
                v_el = ET.Element(f"{ns_uri}v")
                v_el.text = val_str
                c_el.append(v_el)
                
            row_el.append(c_el)
            
        rows_to_keep.append(row_el)
        
    # 4. Serialize rows
    ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
    ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
    ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')
    
    new_rows_xml = []
    for row in rows_to_keep:
        row_str = ET.tostring(row, encoding='utf-8').decode('utf-8')
        # strip namespaces from individual elements so we don't repeat them
        row_str = re.sub(r'\s*xmlns(:\w+)?="[^"]*"', '', row_str)
        new_rows_xml.append(row_str)
        
    new_sheet_data_content = "".join(new_rows_xml)
    
    # 5. Assemble final XML
    final_xml = prefix + sheet_data_open_tag + new_sheet_data_content + end_tag + suffix
    
    # 6. Update dimension ref dynamically to prevent Excel repair errors
    max_row = 3 + len(rows_data)
    final_xml = re.sub(r'ref="A1:([A-Z]+)\d+"', f'ref="A1:\\g<1>{max_row}"', final_xml)
    
    files[sheet_xml_path] = final_xml.encode('utf-8')
    
    out_buffer = io.BytesIO()
    with zipfile.ZipFile(out_buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, content in files.items():
            zout.writestr(name, content)
            
    return out_buffer.getvalue()

def get_pack_contains_gender(gender_val):
    g = str(gender_val or "").strip().lower()
    if "girl" in g:
        return "Girl's"
    elif "boy" in g:
        return "Boy's"
    elif "unisex" in g:
        return "Unisex"
    elif "women" in g:
        return "Women's"
    elif "men" in g:
        return "Men's"
    return "Boy's" # default

def clean_material_value(material_val, fabric_val, category_name, options):
    """
    Cleans and maps the material/fabric from Item Directory to the mastersheet Material options.
    If category is Denim/Jeans, returns 'Denim'.
    If category is Apparel, defaults to 'Cotton' if not matched.
    """
    m = str(material_val or "").strip().upper()
    f = str(fabric_val or "").strip().upper()
    
    # Filter out placeholders
    if m in ["", "NONE", "NAN", "(NIL)", "NIL"]:
        m = ""
    if f in ["", "NONE", "NAN", "(NIL)", "NIL"]:
        f = ""
        
    cat = str(category_name).lower().strip()
    is_denim = "denim" in cat or "jeans" in cat or "jean" in cat
    
    # If it is Denim/Jeans, default to Denim
    if is_denim:
        for opt in options:
            if opt.lower() == "denim":
                return opt
        return "Denim"
        
    # Check if category is Apparel
    is_footwear = any(kw in cat for kw in ["shoes", "sandal", "footwear", "slipper", "clog", "moulds", "sole"])
    
    # Try to find exact match in options
    for val in [m, f]:
        if not val:
            continue
        # Direct match (case-insensitive)
        for opt in options:
            if opt.lower() == val.lower():
                return opt
                
        # Substring match
        for opt in options:
            if opt.lower() in val.lower():
                if "cotton" in opt.lower() and "cotton" in val.lower():
                    return opt
                if not is_footwear or opt.lower() not in ["cotton"]:
                    return opt
                    
    # Fallbacks
    if not is_footwear:
        for opt in options:
            if opt.lower() == "cotton":
                return opt
        return "Cotton"
        
    for val in [m, f]:
        if "eva" in val.lower():
            for opt in options:
                if opt.lower() == "eva":
                    return opt
            return "EVA"
            
    for opt in options:
        if opt.lower() == "synthetic":
            return opt
            
    return "Synthetic"

def map_category_to_product_type(item_cat, item_subcat, options):
    """
    Maps CATEGORY and SUB CATEGORY from the Item Directory to a valid option in Kids Product Type.
    """
    cat = str(item_cat or "").strip().upper()
    subcat = str(item_subcat or "").strip().upper()
    
    # If category is Denim, use subcategory
    target = subcat if "DENIM" in cat else cat
    if not target:
        target = subcat or cat
        
    target = target.upper()
    
    # 1. Run explicit overrides first to prevent substring collisions (e.g. Tshirt matching Shirts)
    t_lower = target.lower()
    if "t-shirt" in t_lower or "tshirt" in t_lower or "t shirt" in t_lower:
        if "polo" in t_lower:
            for opt in options:
                if "polo" in opt.lower():
                    return opt
        for opt in options:
            if opt.lower() == "t-shirts":
                return opt
    if "jeans" in t_lower or "jean" in t_lower:
        for opt in options:
            if opt.lower() == "jeans":
                return opt
    if "trouser" in t_lower:
        for opt in options:
            if opt.lower() == "trousers":
                return opt
    if "short" in t_lower:
        for opt in options:
            if opt.lower() == "shorts":
                return opt
    if "dress" in t_lower:
        for opt in options:
            if "dress" in opt.lower():
                return opt

    # 2. Look for exact match (case-insensitive)
    for opt in options:
        if opt.upper() == target:
            return opt
            
    # 3. Fuzzy match
    def clean_str(s):
        s = s.lower().replace("-", "").replace(" ", "").replace("_", "")
        if s.endswith("s"):
            s = s[:-1]
        return s
        
    target_clean = clean_str(target)
    for opt in options:
        opt_clean = clean_str(opt)
        if opt_clean == target_clean or opt_clean in target_clean or target_clean in opt_clean:
            return opt
            
    # 4. General keyword substring matches
    for opt in options:
        if opt.lower() in t_lower or t_lower in opt.lower():
            return opt
            
    return target.title()

def generate_nykaa_template(db: Session, job: ProcessingJob, logger: EngineLogger):
    logger.log("info", f"Starting Listing Generation Job: {job.job_id}")

    # 1. Fetch Active Files from Database
    item_dir_file = db.query(DbFile).filter(DbFile.file_type == "item_directory").order_by(DbFile.uploaded_at.desc()).first()
    content_sheet_file = db.query(DbFile).filter(DbFile.file_type == "content_sheet").order_by(DbFile.uploaded_at.desc()).first()

    if not item_dir_file:
        logger.log("error", "No Item Directory found in database.")
        raise ValueError("Please upload an Item Directory.")
    if not content_sheet_file:
        logger.log("error", "No Content Sheet found in database.")
        raise ValueError("Please upload a Content Sheet.")

    # 2. Load Category Configuration (always default to Kids Clothing Core if job.category is empty or Kids Clothing Core)
    cat_name = job.category if job.category else "Kids Clothing Core"
    cat_config = db.query(CategoryConfig).filter(CategoryConfig.category_name == cat_name).first()
    if not cat_config or not cat_config.template_file_id:
        # Fallback to search any CategoryConfig with template_file_id
        cat_config = db.query(CategoryConfig).filter(CategoryConfig.template_file_id != None).first()
        
    if not cat_config or not cat_config.template_file_id:
        logger.log("error", f"No template config found for category '{cat_name}'.")
        raise ValueError(f"Category template is not configured. Please upload a template in Upload Center.")

    template_file = db.query(DbFile).filter(DbFile.id == cat_config.template_file_id).first()
    if not template_file:
        logger.log("error", "The template file was not found in the database.")
        raise ValueError("Template file missing in database.")

    # 3. Read template bytes from DB
    template_bytes = decode_and_decompress(template_file.content_b64)

    # 4. Extract sheet name, column mapping, and dropdown options via openpyxl read-only (in-memory)
    logger.log("info", "Extracting sheet metadata, column mapping, and dropdown validation lists from template in-memory...")
    target_sheet_name, col_mapping, size_columns, age_options, kids_product_type_options, material_options = read_col_mapping_via_openpyxl_read_only(template_bytes, cat_name, logger)
    logger.log("info", f"Resolved category sheet name: '{target_sheet_name}'")
    logger.log("info", f"Size dimension columns: {size_columns}")

    # 5. Load Item Directory & Content Sheet
    logger.log("info", "Loading Item Directory...")
    item_bytes = decode_and_decompress(item_dir_file.content_b64)
    item_df = pd.read_excel(io.BytesIO(item_bytes), sheet_name=0)
    logger.log("success", f"Loaded Item Directory: {len(item_df)} rows.")

    # 6. Load Content Sheet
    logger.log("info", "Loading Content Sheet...")
    content_bytes = decode_and_decompress(content_sheet_file.content_b64)
    content_xl = pd.ExcelFile(io.BytesIO(content_bytes))
    content_sheet_name = "MarketplaceD2C" if "MarketplaceD2C" in content_xl.sheet_names else content_xl.sheet_names[0]
    content_df = pd.read_excel(io.BytesIO(content_bytes), sheet_name=content_sheet_name)
    logger.log("success", f"Loaded Content Sheet '{content_sheet_name}': {len(content_df)} rows.")

    # 6. Parse input codes
    raw_codes = job.input_codes or ""
    color_codes = [c.strip() for c in re.split(r'[\n\r,;]+', raw_codes) if c.strip()]
    logger.log("info", f"Parsing target Item Colors: {color_codes}")
    if not color_codes:
        raise ValueError("No Item Color codes provided for generation.")

    # Normalize columns for matching
    item_df['clean_item_name'] = item_df['ITEM NAME'].astype(str).str.strip().str.upper()
    item_df['clean_color'] = item_df['COLOR'].astype(str).str.strip().str.upper()
    item_df['clean_item_color'] = item_df['Item Color'].astype(str).str.strip().str.upper()
    item_df['norm_color'] = item_df['COLOR'].apply(normalize_color_val)
    item_df['norm_item_color'] = item_df['Item Color'].apply(normalize_color_val)
    
    content_df['clean_item_name'] = content_df['Item Name'].astype(str).str.strip().str.upper()
    content_df['clean_shade_name'] = content_df['SHADE NAME'].astype(str).str.strip().str.upper()
    content_df['norm_shade_name'] = content_df['SHADE NAME'].apply(normalize_color_val)

    # Load apparel size charts
    load_apparel_size_charts()

    # Cache for size mappings from DB
    size_mappings_cache = {}
    def get_size_mappings_for_category(c_name):
        c_name_upper = c_name.upper().strip()
        if c_name_upper not in size_mappings_cache:
            rows = db.query(SizeMapping).filter(SizeMapping.category_name == c_name).all()
            size_mappings_cache[c_name_upper] = {str(sz.brand_size).strip().upper(): sz.measurements for sz in rows}
        return size_mappings_cache[c_name_upper]

    # 7. Build data rows
    job.progress = 30
    db.commit()

    rows_to_generate = []
    missing_mappings_report = []

    for code in color_codes:
        logger.log("info", f"Processing code: '{code}'...")
        parts = code.rsplit("-", 1)
        if len(parts) != 2:
            logger.log("warning", f"Invalid format '{code}'. Expected STYLE-COLOR. Skipping.")
            missing_mappings_report.append({"code": code, "issue": "Invalid format (Expected STYLE-COLOR)"})
            continue

        style_code, color_val = parts[0].strip().upper(), parts[1].strip().upper()
        norm_color_val = normalize_color_val(color_val)

        sku_rows = item_df[
            (item_df['clean_item_name'] == style_code) &
            ((item_df['norm_color'] == norm_color_val) | (item_df['norm_item_color'] == norm_color_val))
        ]
        if sku_rows.empty:
            logger.log("warning", f"No SKUs found for Style: '{style_code}', Color: '{color_val}'")
            missing_mappings_report.append({"code": code, "issue": "Not found in Item Directory"})
            continue

        logger.log("info", f"Found {len(sku_rows)} size variants.")

        # Determine dynamic resolved category from the first matching row in Item Directory
        first_sku = sku_rows.iloc[0]
        raw_category = clean_excel_value(first_sku.get('CATEGORY'))
        raw_subcategory = clean_excel_value(first_sku.get('SUB CATEGORY'))
        
        resolved_category = raw_subcategory if "DENIM" in raw_category.upper() else raw_category
        if not resolved_category:
            resolved_category = raw_subcategory or raw_category or "Tshirt"
            
        resolved_category_title = resolved_category.title()
        logger.log("info", f"Dynamically resolved category for style '{style_code}': '{resolved_category_title}' (Raw Category: '{raw_category}', Subcategory: '{raw_subcategory}')")

        content_row = content_df[
            (content_df['clean_item_name'] == style_code) &
            (content_df['norm_shade_name'] == norm_color_val)
        ]
        if content_row.empty:
            content_row = content_df[content_df['clean_item_name'] == style_code]
            if not content_row.empty:
                logger.log("info", f"Using style-only content for '{style_code}'.")
            else:
                logger.log("warning", f"No content sheet match for '{style_code}'")

        nykaa_title = ""
        nykaa_desc = ""
        image_url = ""
        if not content_row.empty:
            c_first = content_row.iloc[0]
            nykaa_title = clean_excel_value(c_first.get('Nykaa Title'))
            nykaa_desc = clean_excel_value(c_first.get('Description'))
            image_url = clean_excel_value(c_first.get('Product Image'))

        if not nykaa_title:
            nykaa_title = f"{first_sku.get('Brand', 'Purple United Kids')} {resolved_category_title}"
            logger.log("warning", f"Fallback title: '{nykaa_title}'")

        for _, sku_row in sku_rows.iterrows():
            brand_size = str(sku_row.get('SIZE', '')).strip()
            barcode = clean_excel_value(sku_row.get('ITEM CODE'))
            mrp = sku_row.get('MRP', 0)
            hsn = clean_excel_value(sku_row.get('HS CODE'))

            raw_gender = clean_excel_value(sku_row.get('GENDER'))
            gender_mapped = map_dropdown_gender(raw_gender)
            design_group = get_design_code(resolved_category_title, gender_mapped)
            multipack = "Combo" if "set" in resolved_category_title.lower() else "Single"

            cat_lower = resolved_category_title.lower()
            is_footwear = any(kw in cat_lower for kw in ["shoes", "sandal", "footwear", "slipper", "heel", "flat", "sneaker", "boot", "clog"])
            care_inst = "Wipe it with a clean dry cloth." if is_footwear else "Hand Wash"

            # Dynamic Pack Contains logic: "1 + Gender's + Category" (e.g. 1 Boy's T-shirt)
            pack_contains = f"1 {get_pack_contains_gender(raw_gender)} {get_singular_category_item(resolved_category_title)}"

            # Clean and map material/fabric dynamically
            raw_material = clean_excel_value(sku_row.get('MATERIAL'))
            raw_fabric = clean_excel_value(sku_row.get('FABRIC'))
            material_val = clean_material_value(raw_material, raw_fabric, resolved_category_title, material_options)

            # Map Kids Product Type dynamically
            product_type_val = map_category_to_product_type(raw_category, raw_subcategory, kids_product_type_options)

            sku_data = {
                "Vendor SKU Code": barcode,
                "Ean Codes": barcode,
                "Style Code": style_code,
                "Price": mrp,
                "Color": clean_excel_value(sku_row.get('COLOR', color_val)).capitalize(),
                "brand  size": brand_size,
                "Design Code": design_group,
                "Gender": gender_mapped,
                "Product Name": nykaa_title,
                "Description": nykaa_desc,
                "HSN Codes": hsn,
                "Country of Origin": "India" if "DOMESTIC" in str(sku_row.get('IMPORTED/DOMESTIC', '')).upper() else "Vietnam",
                "Brand Name": "Purple United Kids" if "toothless" not in str(sku_row.get('Brand', '')).lower() else "toothless",
                "Manufacturer Name": "Purple United Sales Ltd.",
                "Manufacturer Address": "Kh. No. 55/14 And 55/15, Mundka, Delhi 110041",
                "Multipack Set": multipack,
                "Occasion": "Casual",
                "Care Instruction": care_inst,
                "Ships In Days": 1,
                "Pack Contains": pack_contains,
                "Net Qty": "1N",
                "Material": material_val,
                "Kids Product Type": product_type_val,
                "Pattern": "Solid/Plain",
                "Fit": "Regular",
                "Category Classification Kids": "Westernwear"
            }

            if age_options:
                mapped_age = map_size_to_dropdown_age(brand_size, age_options)
                if mapped_age:
                    sku_data["Age"] = mapped_age

            # Fallback hardcoded defaults from Db config if any matching resolved_category_title
            resolved_config = db.query(CategoryConfig).filter(CategoryConfig.category_name == resolved_category_title).first()
            if resolved_config and resolved_config.hardcoded_values:
                for k, v in resolved_config.hardcoded_values.items():
                    if k not in sku_data or sku_data[k] in ["", None]:
                        sku_data[k] = v

            # Size chart measurements based on resolved category section
            section = get_size_chart_section(resolved_category_title, gender_mapped)
            norm_size = brand_size.upper().strip()

            for col_name in size_columns:
                val = None
                if col_name in ["Length (Inches)", "Length For Body (Inches)"]:
                    length_cm = sku_row.get("LENGTH IN CM")
                    if length_cm is not None and pd.notna(length_cm) and str(length_cm).strip() not in ["", "0", "0.0", "nan", "NaN"]:
                        val = convert_cm_to_inches(length_cm)

                if val is None:
                    raw_val = get_chart_measurement_by_keyword(gender_mapped, section, col_name, brand_size)
                    if raw_val is not None:
                        val = convert_cm_to_inches(raw_val)

                if val is None:
                    size_mappings_dict = get_size_mappings_for_category(resolved_category_title)
                    measurements = size_mappings_dict.get(norm_size)
                    if measurements and col_name in measurements:
                        val = measurements[col_name]
                        try:
                            val = round(float(val), 1)
                        except (ValueError, TypeError):
                            pass

                if val is not None:
                    sku_data[col_name] = val
                else:
                    logger.log("warning", f"No size data for '{col_name}' size '{brand_size}'")
                    if col_name in ["Chest for Garment (Inches)", "Waist for Garment (Inches)"]:
                        missing_mappings_report.append({"code": code, "issue": f"Missing '{col_name}' for size {brand_size}"})

            if image_url:
                sku_data["Front Image"] = image_url

            rows_to_generate.append(sku_data)

    logger.log("info", f"Built {len(rows_to_generate)} SKU rows. Now filling template in-memory via direct ZIP/XML modification...")
    job.progress = 60
    db.commit()

    # 8. Fill data directly into the original template in-memory
    final_bytes = fill_template_via_xml_manipulation(template_bytes, target_sheet_name, col_mapping, rows_to_generate, logger)

    job.progress = 90
    db.commit()

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_filename = f"Nykaa_Populated_{target_sheet_name}_{timestamp_str}.xlsm"
    out_b64 = base64.b64encode(zlib.compress(final_bytes, level=9)).decode("utf-8")

    out_db_file = DbFile(
        file_type="output_file",
        filename=out_filename,
        content_b64=out_b64
    )
    db.add(out_db_file)
    db.flush()

    job.progress = 100
    job.status = "success"
    job.output_filename = out_filename
    job.output_file_id = out_db_file.id
    job.validation_report = {
        "total_sku_rows": len(rows_to_generate),
        "errors": missing_mappings_report
    }
    db.commit()
    logger.log("success", f"Generated {len(rows_to_generate)} SKUs in '{out_filename}'.")
    logger.log("success", "=== NYKAA AUTO LISTING WORKFLOW COMPLETED ===")

